import os
import io
import math
import random
import pandas as pd
from enum import Enum
from datetime import datetime
from pydantic import BaseModel
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

from app.db.conexion import conectar_base
from app.utils.generador_pdf import generar_documento
from app.schemas.usuarios import AlumnoCrear, PersonalCrear
from app.utils.seguridad import obtener_hash_contrasena
from app.utils.notificaciones import (
    enviar_correo_rechazo,
    enviar_correo_documento,
    generar_password_seguro,
    enviar_correo_credenciales,
    enviar_correo_desbloqueo
)

router = APIRouter()

# ==============================================================================
# SCHEMAS PYDANTIC
# ==============================================================================

class CambioEstadoUsuario(BaseModel):
    numero_cuenta: str
    estado_activo: bool

class CambioRol(BaseModel):
    numero_cuenta: str
    nuevo_rol: str

class OpcionesEstadoDocumento(str, Enum):
    APROBADO = "APROBADO"
    RECHAZADO = "RECHAZADO"

class ActualizarEstadoDocumento(BaseModel):
    id_admin: int 
    estado: OpcionesEstadoDocumento
    comentario: str | None = None

class RechazarSolicitud(BaseModel):
    id_admin: int
    motivo: str

class GenerarDocumentoRequest(BaseModel):
    id_admin: int

class AprobarEstacionamiento(BaseModel):
    id_admin: int
    comentario: str | None = "Solicitud de estacionamiento aprobada."

class AprobarLocker(BaseModel):
    id_admin: int
    id_locker: int
    comentario: str | None = "Solicitud aprobada y locker asignado."

class GuardiaCrear(BaseModel):
    nombre: str
    apellidos: str
    correo_electronico: str

class ApelacionBloqueo(BaseModel):
    id_admin: int
    comentario: str   


# ==============================================================================
# GESTIÓN Y REVISIÓN DE SOLICITUDES
# ==============================================================================

@router.get("/solicitudes/{id_solicitud}/detalle", tags=["Administrador / Personal"], summary="Obtener documentos de una solicitud completa para revisión")
def obtener_detalle_solicitud(id_solicitud: int):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT
                s.id_solicitud,
                s.tipo_tramite,
                s.estado,
                a.numero_cuenta,
                a.nombre || ' ' || a.apellidos AS nombre_completo,
                s.fecha_solicitud,
                ds.id_documento,
                ds.id_tipo_documento,
                td.nombre_tipo_documento,
                ds.archivo_path,
                ds.estado AS estado_documento,
                ds.comentario
            FROM solicitud s
            JOIN alumno a ON s.id_alumno = a.id_alumno
            LEFT JOIN documentos_solicitud ds ON s.id_solicitud = ds.id_solicitud
            LEFT JOIN tipo_documento td ON ds.id_tipo_documento = td.id_tipo_documento
            WHERE s.id_solicitud = %s
            ORDER BY ds.id_tipo_documento
        """, (id_solicitud,))

        filas = cursor.fetchall()
        cursor.close()
        conexion.close()

        if not filas:
            raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

        primera = filas[0]
        detalle = {
            "id_solicitud": primera[0],
            "tipo_tramite": primera[1],
            "estado": primera[2],
            "numero_cuenta": primera[3],
            "nombre_completo": primera[4],
            "fecha_solicitud": primera[5],
            "documentos": []
        }

        for fila in filas:
            if fila[6] is not None:
                detalle["documentos"].append({
                    "id_documento": fila[6],
                    "id_tipo_documento": fila[7],
                    "nombre_tipo_documento": fila[8],
                    "archivo": fila[9],
                    "estado_documento": fila[10],
                    "comentario_admin": fila[11]
                })

        return detalle

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.get("/solicitudes/", tags=["Administrador / Personal"], summary="Obtener todas las solicitudes de los alumnos")
def obtener_todas_las_solicitudes(
    tipo_tramite: str | None = None, 
    estado: str | None = None, 
    fecha: str | None = None,
    busqueda: str | None = None,
    page: int = 1
):
    registros_por_pagina = 20
    offset = (page - 1) * registros_por_pagina
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        query_base = """
            FROM solicitud s
            JOIN alumno a ON s.id_alumno = a.id_alumno
            WHERE s.estado != 'DATOS_INCOMPLETOS'
        """
        parametros = []

        if tipo_tramite:
            query_base += " AND s.tipo_tramite = %s"
            parametros.append(tipo_tramite.lower())
            
        if estado:
            query_base += " AND s.estado = %s"
            parametros.append(estado.upper())
            
        if fecha:
            query_base += " AND DATE(s.fecha_solicitud) = %s"
            parametros.append(fecha)
            
        if busqueda:
            query_base += " AND (a.numero_cuenta ILIKE %s OR a.nombre ILIKE %s OR a.apellidos ILIKE %s)"
            busqueda_like = f"%{busqueda}%"
            parametros.extend([busqueda_like, busqueda_like, busqueda_like])

        query_count = f"SELECT COUNT(*) {query_base}"
        cursor.execute(query_count, tuple(parametros))
        total_registros = cursor.fetchone()[0]

        query_paginada = f"""
            SELECT 
                s.id_solicitud,
                a.numero_cuenta,
                a.nombre,
                a.apellidos,
                s.tipo_tramite,
                s.fecha_solicitud,
                s.estado
            {query_base}
            ORDER BY s.fecha_solicitud DESC
            LIMIT %s OFFSET %s
        """
        parametros.extend([registros_por_pagina, offset])
        cursor.execute(query_paginada, tuple(parametros))
        filas = cursor.fetchall()
        
        solicitudes = []
        for fila in filas:
            solicitudes.append({
                "id_solicitud": fila[0],
                "folio": f"FOL-{fila[0]:04d}", 
                "numero_cuenta": fila[1],
                "nombre_completo": f"{fila[2]} {fila[3]}",
                "tipo_tramite": fila[4],
                "fecha_solicitud": fila[5],
                "estado": fila[6]
            })

        cursor.close()
        conexion.close()
        total_paginas = math.ceil(total_registros / registros_por_pagina) if total_registros > 0 else 1

        return {
            "pagina_actual": page,
            "registros_por_pagina": registros_por_pagina,
            "total_registros": total_registros,
            "total_paginas": total_paginas,
            "resultados": solicitudes
        }

    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.put("/solicitudes/{id_solicitud}/documentos/{id_documento}", tags=["Administrador / Personal"], summary="Evaluar documento individualmente de cada solicitud")
def evaluar_documento_individual(
    id_solicitud: int,
    id_documento: int,
    datos: ActualizarEstadoDocumento
):
    estado_texto = datos.estado.value 
    comentario_final = datos.comentario

    if estado_texto == "RECHAZADO":
        if not comentario_final or comentario_final.strip() == "":
            raise HTTPException(status_code=400, detail="Debe proporcionar un comentario explicando por qué se rechaza el documento.")
    elif estado_texto == "APROBADO":
        comentario_final = None

    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT id_documento FROM documentos_solicitud
            WHERE id_solicitud = %s AND id_documento = %s
        """, (id_solicitud, id_documento))

        if not cursor.fetchone():
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Documento no encontrado o no pertenece a la solicitud.")

        cursor.execute("""
            UPDATE documentos_solicitud
            SET estado = %s, comentario = %s
            WHERE id_solicitud = %s AND id_documento = %s
        """, (estado_texto, comentario_final, id_solicitud, id_documento))

        cursor.execute("""
            UPDATE solicitud
            SET revisado_por = %s, fecha_revision = CURRENT_TIMESTAMP
            WHERE id_solicitud = %s AND estado IN ('PENDIENTE', 'REPOSICION')
        """, (datos.id_admin, id_solicitud))

        conexion.commit()
        cursor.close()
        conexion.close()

        return {
            "mensaje": "Documento evaluado correctamente.",
            "id_solicitud": id_solicitud,
            "id_documento": id_documento,
            "nuevo_estado": estado_texto,
            "comentario_registrado": comentario_final
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.post("/solicitudes/{id_solicitud}/rechazar", tags=["Administrador / Personal"], summary="Rechazar solicitud completa y notificar via correo electronico el motivo")
def rechazar_solicitud(id_solicitud: int, datos: RechazarSolicitud):
    if not datos.motivo or datos.motivo.strip() == "":
        raise HTTPException(status_code=400, detail="El motivo de rechazo es obligatorio.")

    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT s.estado, s.tipo_tramite, a.nombre, a.apellidos, a.correo_electronico, a.numero_cuenta
            FROM solicitud s
            JOIN alumno a ON s.id_alumno = a.id_alumno
            WHERE s.id_solicitud = %s
        """, (id_solicitud,))
        solicitud_actual = cursor.fetchone()

        if not solicitud_actual:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Solicitud no encontrada")

        estado_anterior = solicitud_actual[0]
        
        if estado_anterior == "DOCUMENTACION_INCORRECTA":
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="Esta solicitud ya ha sido rechazada anteriormente.")
            
        if estado_anterior == "APROBADA":
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="Esta solicitud ya está aprobada, no se puede rechazar.")

        tramite = solicitud_actual[1]
        nombre_completo = f"{solicitud_actual[2]} {solicitud_actual[3]}" 
        correo_alumno = solicitud_actual[4] 
        num_cuenta_alumno = solicitud_actual[5]
        nuevo_estado = "DOCUMENTACION_INCORRECTA"

        cursor.execute("""
            UPDATE solicitud
            SET estado = %s, revisado_por = %s, fecha_revision = CURRENT_TIMESTAMP
            WHERE id_solicitud = %s
        """, (nuevo_estado, datos.id_admin, id_solicitud))

        cursor.execute("""
            INSERT INTO historial_estados (id_solicitud, estado_anterior, estado_nuevo, id_admin, comentario)
            VALUES (%s, %s, %s, %s, %s)
        """, (id_solicitud, estado_anterior, nuevo_estado, datos.id_admin, datos.motivo))
        
        cursor.execute("""
            INSERT INTO notificaciones (numero_cuenta, rol_destino, titulo, mensaje)
            VALUES (%s, 'ALUMNO', 'Trámite Rechazado', %s)
        """, (num_cuenta_alumno, f"Tu solicitud de {tramite} fue rechazada. Revisa tu correo o el sistema para corregir."))

        conexion.commit()
        correo_enviado = enviar_correo_rechazo(correo_alumno, nombre_completo, tramite, datos.motivo)

        cursor.close()
        conexion.close()

        mensaje_respuesta = "Solicitud rechazada y auditoría registrada exitosamente."
        if not correo_enviado:
            mensaje_respuesta += " (Advertencia: Guardado en BD, pero falló el envío de correo)."

        return {
            "mensaje": mensaje_respuesta,
            "id_solicitud": id_solicitud,
            "estado": nuevo_estado,
            "auditoria": "Guardada en historial_estados",
            "correo_notificado": correo_alumno
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.post("/solicitudes/{id_solicitud}/aprobar-locker", tags=["Administrador / Personal"], summary="Aprobar solicitud de locker y asignarle uno al alumno")
def aprobar_solicitud_locker(id_solicitud: int, datos: AprobarLocker):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT s.id_solicitud, s.estado, s.tipo_tramite, a.numero_cuenta
            FROM solicitud s JOIN alumno a ON s.id_alumno = a.id_alumno
            WHERE s.id_solicitud = %s;
        """, (id_solicitud,))
        solicitud = cursor.fetchone()

        if not solicitud:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

        estado_actual, tipo_tramite, numero_cuenta = solicitud[1], solicitud[2], solicitud[3]

        if tipo_tramite.lower() != "locker":
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="Esta solicitud no es de tipo locker.")

        if estado_actual not in ["PENDIENTE", "EN_REVISION", "REPOSICION"]:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail=f"La solicitud no se puede aprobar porque está en estado {estado_actual}.")

        cursor.execute("SELECT COUNT(*) FROM documentos_solicitud WHERE id_solicitud = %s AND estado != 'APROBADO';", (id_solicitud,))
        if cursor.fetchone()[0] > 0:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="No se puede aprobar la solicitud porque aún hay documentos sin aprobar.")

        cursor.execute("SELECT id_locker, codigo_locker, ubicacion, estado FROM locker WHERE id_locker = %s;", (datos.id_locker,))
        locker = cursor.fetchone()

        if not locker:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Locker no encontrado.")

        if locker[3] != "DISPONIBLE":
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail=f"El locker no está disponible. Estado actual: {locker[3]}.")

        cursor.execute("UPDATE solicitud SET estado = 'APROBADA', revisado_por = %s, fecha_revision = CURRENT_TIMESTAMP WHERE id_solicitud = %s;", (datos.id_admin, id_solicitud))
        cursor.execute("INSERT INTO asignacion (id_solicitud, id_locker, estado) VALUES (%s, %s, 'ACTIVA') RETURNING id_asignacion;", (id_solicitud, datos.id_locker))
        id_asignacion = cursor.fetchone()[0]

        cursor.execute("UPDATE locker SET estado = 'OCUPADO' WHERE id_locker = %s;", (datos.id_locker,))
        cursor.execute("INSERT INTO historial_estados (id_solicitud, estado_anterior, estado_nuevo, id_admin, comentario) VALUES (%s, %s, %s, %s, %s);", (id_solicitud, estado_actual, "APROBADA", datos.id_admin, datos.comentario))
        cursor.execute("INSERT INTO notificaciones (numero_cuenta, rol_destino, titulo, mensaje) VALUES (%s, 'ALUMNO', 'Solicitud de locker aprobada', %s);", (numero_cuenta, f"Tu solicitud de locker fue aprobada. Se te asignó el locker {locker[1]} ubicado en {locker[2]}."))

        conexion.commit()
        cursor.close()
        conexion.close()

        return {
            "mensaje": "Solicitud de locker aprobada correctamente.",
            "id_solicitud": id_solicitud,
            "id_asignacion": id_asignacion,
            "locker_asignado": {
                "id_locker": locker[0],
                "codigo_locker": locker[1],
                "ubicacion": locker[2],
                "estado": "OCUPADO"
            },
            "estado_solicitud": "APROBADA"
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.post("/solicitudes/{id_solicitud}/aprobar-estacionamiento", tags=["Administrador / Personal"], summary="Aprobar solicitud de estacionamiento")
def aprobar_solicitud_estacionamiento(id_solicitud: int, datos: AprobarEstacionamiento):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT s.id_solicitud, s.estado, s.tipo_tramite, a.numero_cuenta
            FROM solicitud s JOIN alumno a ON s.id_alumno = a.id_alumno
            WHERE s.id_solicitud = %s;
        """, (id_solicitud,))
        solicitud = cursor.fetchone()

        if not solicitud:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

        estado_actual, tipo_tramite, numero_cuenta = solicitud[1], solicitud[2], solicitud[3]

        if tipo_tramite.lower() != "estacionamiento":
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="Esta solicitud no es de tipo estacionamiento.")

        if estado_actual not in ["PENDIENTE", "EN_REVISION", "REPOSICION"]:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail=f"La solicitud no se puede aprobar porque está en estado {estado_actual}.")

        cursor.execute("SELECT COUNT(*) FROM documentos_solicitud WHERE id_solicitud = %s AND estado != 'APROBADO';", (id_solicitud,))
        if cursor.fetchone()[0] > 0:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="No se puede aprobar la solicitud porque aún hay documentos sin aprobar.")

        cursor.execute("UPDATE solicitud SET estado = 'APROBADA', revisado_por = %s, fecha_revision = CURRENT_TIMESTAMP WHERE id_solicitud = %s;", (datos.id_admin, id_solicitud))
        cursor.execute("INSERT INTO asignacion (id_solicitud, id_locker, estado) VALUES (%s, NULL, 'ACTIVA') RETURNING id_asignacion;", (id_solicitud,))
        id_asignacion = cursor.fetchone()[0]
        cursor.execute("INSERT INTO historial_estados (id_solicitud, estado_anterior, estado_nuevo, id_admin, comentario) VALUES (%s, %s, %s, %s, %s);", (id_solicitud, estado_actual, "APROBADA", datos.id_admin, datos.comentario))
        cursor.execute("INSERT INTO notificaciones (numero_cuenta, rol_destino, titulo, mensaje) VALUES (%s, 'ALUMNO', 'Solicitud de estacionamiento aprobada', %s);", (numero_cuenta, "Tu solicitud de estacionamiento fue aprobada correctamente. Estás listo para generar tu tarjetón."))

        conexion.commit()
        cursor.close()
        conexion.close()

        return {
            "mensaje": "Solicitud de estacionamiento aprobada correctamente.",
            "id_solicitud": id_solicitud,
            "id_asignacion": id_asignacion,
            "estado_solicitud": "APROBADA"
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.post("/solicitudes/{id_solicitud}/aceptar", tags=["Administrador / Personal"], summary="Aceptar solicitud completa, notificar via correo electronico, generar constancia o tarjeton con QR y calcular caducidad semestral")
def aceptar_solicitud_y_generar_documento(id_solicitud: int, datos: GenerarDocumentoRequest):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT a.id_asignacion, s.id_solicitud, s.tipo_tramite, s.estado, al.id_alumno, al.nombre, al.apellidos, al.numero_cuenta, al.correo_electronico, l.codigo_locker, l.ubicacion
            FROM asignacion a
            JOIN solicitud s ON a.id_solicitud = s.id_solicitud
            JOIN alumno al ON s.id_alumno = al.id_alumno
            LEFT JOIN locker l ON a.id_locker = l.id_locker
            WHERE s.id_solicitud = %s AND a.estado = 'ACTIVA';
        """, (id_solicitud,))
        
        info = cursor.fetchone()
        if not info:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="No se encontró una asignación activa para esta solicitud. Debes aprobar el recurso primero.")
            
        id_asignacion, id_sol_bd, tipo_tramite, estado_solicitud, id_alumno, nombre_al, apellidos_al, num_cuenta, correo_alumno, cod_locker, ubi_locker = info
        nombre_completo = f"{nombre_al} {apellidos_al}"
        
        cod_locker = cod_locker if cod_locker else "N/A"
        ubi_locker = ubi_locker if ubi_locker else "N/A"

        if estado_solicitud != 'APROBADA':
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="La solicitud debe estar APROBADA para generar el documento.")

        cursor.execute("SELECT id_constancia FROM constancia WHERE id_asignacion = %s;", (id_asignacion,))
        if cursor.fetchone():
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=400, detail="Ya se generó un documento oficial para esta solicitud.")

        hoy = datetime.now()
        if hoy.month < 7:
            fecha_vigencia_obj = datetime(hoy.year, 7, 1)
        else:
            fecha_vigencia_obj = datetime(hoy.year + 1, 1, 1)
            
        fecha_vigencia_str = fecha_vigencia_obj.strftime('%Y-%m-%d')
        folio = f"{tipo_tramite[:3].upper()}-{id_asignacion}-{datetime.now().strftime('%m%d')}"

        cursor.execute("""
            INSERT INTO constancia (id_asignacion, folio, vigencia, documento_path)
            VALUES (%s, %s, %s, %s)
            RETURNING qr_token, id_constancia;
        """, (id_asignacion, folio, fecha_vigencia_str, 'pendiente.pdf'))
        
        resultado_insert = cursor.fetchone()
        qr_token = str(resultado_insert[0])
        
        nombre_archivo_pdf = generar_documento(
            folio=folio,
            token_qr=qr_token,
            nombre_alumno=nombre_completo,
            cuenta_alumno=num_cuenta,
            tipo_tramite=tipo_tramite,
            vigencia=fecha_vigencia_str,
            codigo_locker=cod_locker,
            ubicacion_locker=ubi_locker
        )
    
        cursor.execute("""
            UPDATE constancia SET documento_path = %s WHERE id_asignacion = %s;
        """, (nombre_archivo_pdf, id_asignacion))

        cursor.execute("""
            INSERT INTO historial_estados (id_solicitud, estado_anterior, estado_nuevo, id_admin, comentario)
            VALUES (%s, 'APROBADA', 'DOCUMENTO_GENERADO', %s, %s)
        """, (id_solicitud, datos.id_admin, f"Se generó el documento con folio {folio} (Vigencia al {fecha_vigencia_str})."))
        
        correo_enviado = enviar_correo_documento(correo_alumno, nombre_completo, tipo_tramite, nombre_archivo_pdf)
        
        conexion.commit()
        cursor.close()
        conexion.close()

        mensaje_respuesta = "Solicitud aceptada y documento generado exitosamente."
        if not correo_enviado:
            mensaje_respuesta += " (Pero falló el envío de correo)."

        return {
            "mensaje": mensaje_respuesta,
            "folio": folio,
            "qr_token": qr_token,
            "archivo": nombre_archivo_pdf
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD/Motor: {str(e)}")


@router.put("/solicitudes/{id_solicitud}/en-revision", tags=["Administrador / Personal"], summary="Cambiar el estado de una solicitud a EN REVISION cuando el admin la abre")
def marcar_solicitud_en_revision(id_solicitud: int, id_admin: int):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT estado FROM solicitud WHERE id_solicitud = %s", (id_solicitud,))
        resultado = cursor.fetchone()

        if not resultado:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

        estado_actual = resultado[0]
        
        if estado_actual not in ["PENDIENTE", "REPOSICION"]:
            cursor.close()
            conexion.close()
            return {"mensaje": f"La solicitud ya se encuentra en estado {estado_actual}, no se cambió a EN_REVISION."}

        cursor.execute("""
            UPDATE solicitud 
            SET estado = 'EN_REVISION', revisado_por = %s, fecha_revision = CURRENT_TIMESTAMP
            WHERE id_solicitud = %s
        """, (id_admin, id_solicitud))
        
        cursor.execute("""
            INSERT INTO historial_estados (id_solicitud, estado_anterior, estado_nuevo, id_admin, comentario)
            VALUES (%s, %s, 'EN_REVISION', %s, 'El administrador ha comenzado a revisar los documentos.')
        """, (id_solicitud, estado_actual, id_admin))

        conexion.commit()
        cursor.close()
        conexion.close()

        return {"mensaje": "La solicitud ahora está EN REVISIÓN.", "nuevo_estado": "EN_REVISION"}

    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.put("/solicitudes/{id_solicitud}/cancelar-revision", tags=["Administrador / Personal"], summary="Cancelar la revisión si el admin se sale sin hacer nada")
def cancelar_revision(id_solicitud: int):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT estado FROM solicitud WHERE id_solicitud = %s
        """, (id_solicitud,))
        res = cursor.fetchone()
        
        if not res:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

        if res[0] != 'EN_REVISION':
            cursor.close()
            conexion.close()
            return {"mensaje": "No se hizo nada porque la solicitud no está en revisión."}

        cursor.execute("""
            UPDATE solicitud SET estado = 'PENDIENTE', revisado_por = NULL WHERE id_solicitud = %s
        """, (id_solicitud,))
        
        conexion.commit()
        cursor.close()
        conexion.close()

        return {"mensaje": "La solicitud ha regresado a estado PENDIENTE."}

    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.get("/documentos/descargar/{qr_token}", tags=["Administrador / Personal", "Alumno"], summary="Descargar el PDF de un trámite mediante el qr_token")
def descargar_documento(qr_token: str):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")
        
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT documento_path, estado FROM constancia WHERE qr_token = %s::uuid;
        """, (qr_token,))
        
        resultado = cursor.fetchone()
        cursor.close()
        conexion.close()
        
        if not resultado:
            raise HTTPException(status_code=404, detail="Documento no encontrado o código inválido.")
            
        archivo_path = resultado[0]
        estado = resultado[1]
        
        if estado != 'VIGENTE':
            raise HTTPException(status_code=400, detail=f"No se puede descargar. El documento está {estado}.")
             
        ruta_completa = os.path.join("uploads", archivo_path)
        
        if not os.path.exists(ruta_completa):
            raise HTTPException(status_code=404, detail="El archivo físico ya no existe en el servidor.")
            
        return FileResponse(path=ruta_completa, filename=archivo_path, media_type='application/pdf')

    except HTTPException:
        raise
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de token inválido.")
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


@router.get("/solicitudes/documentos/{id_documento}/visualizar", tags=["Administrador / Personal", "Alumno"], summary="Previsualizar un documento y el nombre de forma segura sin exponer datos del alumno publicamente")
def ver_documento_subido(id_documento: int):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")
        
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT archivo_path FROM documentos_solicitud WHERE id_documento = %s", (id_documento,))
        resultado = cursor.fetchone()
        
        cursor.close()
        conexion.close()
        
        if not resultado:
            raise HTTPException(status_code=404, detail="Documento no encontrado en la base de datos.")
            
        archivo_path = resultado[0]
        ruta_completa = os.path.join("uploads", archivo_path)
        
        if not os.path.exists(ruta_completa):
            raise HTTPException(status_code=404, detail="El archivo físico ya no existe en el servidor.")
            
        extension = archivo_path.lower().split('.')[-1]
        media_type = "application/pdf" if extension == "pdf" else f"image/{extension}"
            
        return FileResponse(
            path=ruta_completa, 
            media_type=media_type, 
            filename=archivo_path,
            content_disposition_type="inline" 
        )

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


# ==============================================================================
# ADMINISTRACIÓN DE USUARIOS
# ==============================================================================

@router.post("/admin/usuarios/alumno", tags=["Administrador / Personal"], summary="Crear cuenta del Alumno, autogenerando contraseña y enviando correo")
def crear_cuenta_alumno(datos: AlumnoCrear):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    password_plana = generar_password_seguro()
    hash_password = obtener_hash_contrasena(password_plana)

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO alumno (numero_cuenta, nombre, apellidos, carrera, correo_electronico, contrasena_hash, estado_activo)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id_alumno;
        """, (datos.numero_cuenta, datos.nombre, datos.apellidos, datos.carrera, datos.correo_electronico, hash_password))
        
        id_nuevo = cursor.fetchone()[0]
        conexion.commit()
        
        nombre_completo = f"{datos.nombre} {datos.apellidos}"
        enviar_correo_credenciales(datos.correo_electronico, nombre_completo, datos.numero_cuenta, password_plana, "Alumno")
        
        cursor.close()
        conexion.close()
        return {"mensaje": "Cuenta de alumno creada exitosamente. La contraseña fue enviada a su correo electrónico.", "id_alumno": id_nuevo}

    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        if "unique constraint" in str(e).lower():
            raise HTTPException(status_code=400, detail="El número de cuenta ya está registrado.")
        raise HTTPException(status_code=500, detail=f"Error en BD/Servidor: {str(e)}")


@router.post("/admin/usuarios/personal", tags=["Administrador / Personal"], summary="Crear una cuenta unicamente del Administrador o Personal")
def crear_cuenta_personal(datos: PersonalCrear):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    hash_password = obtener_hash_contrasena(datos.contrasena)

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO admin (numero_cuenta, nombre, apellidos, correo_electronico, contrasena_hash, rol, estado_activo)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id_admin;
        """, (datos.numero_cuenta, datos.nombre, datos.apellidos, datos.correo_electronico, hash_password, datos.rol.upper()))
        
        id_nuevo = cursor.fetchone()[0]
        conexion.commit()

        cursor.close()
        conexion.close()

        return {"mensaje": f"Cuenta de {datos.rol.lower()} creada exitosamente.", "id_admin": id_nuevo}

    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        if "unique constraint" in str(e).lower():
            raise HTTPException(status_code=400, detail="El número de cuenta ya está registrado.")
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.post("/admin/usuarios/guardia", tags=["Administrador / Personal"], summary="Crear cuenta del Guardia autogenerando cuenta y contraseña")
def crear_cuenta_guardia(datos: GuardiaCrear):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    numero_cuenta_guardia = f"900{random.randint(1000, 9999)}"
    password_plana = generar_password_seguro()
    hash_password = obtener_hash_contrasena(password_plana)

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            INSERT INTO admin (numero_cuenta, nombre, apellidos, correo_electronico, contrasena_hash, rol, estado_activo)
            VALUES (%s, %s, %s, %s, %s, 'VIGILANTE', TRUE)
            RETURNING id_admin;
        """, (numero_cuenta_guardia, datos.nombre, datos.apellidos, datos.correo_electronico, hash_password))
        
        id_nuevo = cursor.fetchone()[0]
        conexion.commit()
        
        nombre_completo = f"{datos.nombre} {datos.apellidos}"
        enviar_correo_credenciales(datos.correo_electronico, nombre_completo, numero_cuenta_guardia, password_plana, "Guardia de Seguridad")
        
        cursor.close()
        conexion.close()
        return {
            "mensaje": f"Guardia creado exitosamente. Se le envió su número de cuenta ({numero_cuenta_guardia}) y contraseña a su correo.", 
            "id_admin": id_nuevo,
            "numero_cuenta_generado": numero_cuenta_guardia
        }

    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD/Servidor: {str(e)}")


@router.get("/admin/usuarios", tags=["Administrador / Personal"], summary="Obtener lista de usuarios dentro del sistema separados con filtros y mediante bloques de datos de 20 en 20")
def listar_usuarios(
    rol: str | None = None, 
    estado_activo: bool | None = None, 
    busqueda: str | None = None, 
    page: int = 1
):
    registros_por_pagina = 20
    offset = (page - 1) * registros_por_pagina

    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        query_base = """
            WITH UsuariosCompletos AS (
                SELECT id_alumno as id, numero_cuenta, nombre, apellidos, correo_electronico, estado_activo, 'ALUMNO' as rol 
                FROM alumno
                UNION ALL
                SELECT id_admin as id, numero_cuenta, nombre, apellidos, correo_electronico, estado_activo, rol 
                FROM admin
            )
            SELECT id, numero_cuenta, nombre, apellidos, correo_electronico, estado_activo, rol
            FROM UsuariosCompletos
            WHERE 1=1
        """
        filtros = []

        if rol:
            query_base += " AND rol = %s"
            filtros.append(rol.upper())
            
        if estado_activo is not None:
            query_base += " AND estado_activo = %s"
            filtros.append(estado_activo)
            
        if busqueda:
            query_base += " AND (numero_cuenta ILIKE %s OR nombre ILIKE %s OR apellidos ILIKE %s)"
            busqueda_like = f"%{busqueda}%"
            filtros.extend([busqueda_like, busqueda_like, busqueda_like])

        query_count = f"SELECT COUNT(*) FROM ({query_base}) AS conteo"
        cursor.execute(query_count, tuple(filtros))
        total_registros = cursor.fetchone()[0]

        query_paginada = query_base + " ORDER BY rol, apellidos LIMIT %s OFFSET %s"
        filtros.extend([registros_por_pagina, offset])
        
        cursor.execute(query_paginada, tuple(filtros))
        filas = cursor.fetchall()
        
        resultados = []
        for fila in filas:
            resultados.append({
                "id_usuario": fila[0],
                "numero_cuenta": fila[1],
                "nombre_completo": f"{fila[2]} {fila[3]}",
                "correo_electronico": fila[4],
                "estado_activo": fila[5],
                "rol": fila[6]
            })

        cursor.close()
        conexion.close()
        total_paginas = math.ceil(total_registros / registros_por_pagina) if total_registros > 0 else 1

        return {
            "pagina_actual": page,
            "registros_por_pagina": registros_por_pagina,
            "total_registros": total_registros,
            "total_paginas": total_paginas,
            "resultados": resultados
        }

    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")
    

@router.put("/admin/usuarios/estado", tags=["Administrador / Personal"], summary="Activar, desactivar o bloquear una cuenta")
def cambiar_estado_usuario(datos: CambioEstadoUsuario):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        
        cursor.execute("""
            UPDATE alumno SET estado_activo = %s WHERE numero_cuenta = %s RETURNING id_alumno;
        """, (datos.estado_activo, datos.numero_cuenta))
        resultado_alumno = cursor.fetchone()

        if resultado_alumno:
            conexion.commit()
            cursor.close()
            conexion.close()
            return {"mensaje": f"Estado del alumno actualizado a: {'Activo' if datos.estado_activo else 'Inactivo'}"}

        cursor.execute("""
            UPDATE admin SET estado_activo = %s WHERE numero_cuenta = %s RETURNING id_admin;
        """, (datos.estado_activo, datos.numero_cuenta))
        resultado_personal = cursor.fetchone()

        if resultado_personal:
            conexion.commit()
            cursor.close()
            conexion.close()
            return {"mensaje": f"Estado del personal actualizado a: {'Activo' if datos.estado_activo else 'Inactivo'}"}

        cursor.close()
        conexion.close()
        raise HTTPException(status_code=404, detail="El número de cuenta no existe en el sistema.")

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")
    

@router.put("/admin/usuarios/rol", tags=["Administrador / Personal"], summary="Cambiar el rol de cualquier usuario")
def cambiar_rol_personal(datos: CambioRol):
    rol_upper = datos.nuevo_rol.upper()
    if rol_upper not in ["ADMIN", "REVISOR", "VIGILANTE"]:
        raise HTTPException(status_code=400, detail="Rol inválido. Debe ser ADMIN, REVISOR o VIGILANTE.")

    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            UPDATE admin SET rol = %s WHERE numero_cuenta = %s RETURNING id_admin;
        """, (rol_upper, datos.numero_cuenta))
        
        resultado = cursor.fetchone()

        if not resultado:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="El usuario no existe o es un alumno (los alumnos no pueden cambiar de rol).")

        conexion.commit()
        cursor.close()
        conexion.close()
        return {"mensaje": f"Rol actualizado exitosamente a {rol_upper}."}

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


# ==============================================================================
# AUDITORÍA, STRIKES, APELACIONES Y EXCEL
# ==============================================================================

@router.put("/admin/usuarios/{numero_cuenta}/desbloquear-acceso", tags=["Administrador / Personal"], summary="Desbloquear cuenta de alumno y restaurar permisos por apelación")
def desbloquear_acceso_alumno(numero_cuenta: str, datos: ApelacionBloqueo):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        
        cursor.execute("SELECT id_alumno, nombre, apellidos, correo_electronico FROM alumno WHERE numero_cuenta = %s", (numero_cuenta,))
        alumno_bd = cursor.fetchone()
        
        if not alumno_bd:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Alumno no encontrado.")
            
        id_alumno = alumno_bd[0]
        nombre_completo = f"{alumno_bd[1]} {alumno_bd[2]}"
        correo_electronico = alumno_bd[3]

        cursor.execute("UPDATE alumno SET estado_activo = TRUE WHERE id_alumno = %s", (id_alumno,))
        
        cursor.execute("""
            UPDATE asignacion a
            SET estado = 'ACTIVA'
            FROM solicitud s
            WHERE a.id_solicitud = s.id_solicitud 
              AND s.id_alumno = %s 
              AND a.estado = 'BLOQUEADA'
            RETURNING s.id_solicitud
        """, (id_alumno,))
        solicitudes_restauradas = cursor.fetchall()

        cursor.execute("""
            UPDATE constancia c
            SET estado = 'VIGENTE'
            FROM asignacion a, solicitud s
            WHERE c.id_asignacion = a.id_asignacion
              AND a.id_solicitud = s.id_solicitud
              AND s.id_alumno = %s
              AND c.estado = 'BLOQUEADO'
        """, (id_alumno,))
        
        # Eliminar el strike más reciente en acceso_denegado si existe la tabla
        cursor.execute("""
            DELETE FROM acceso_denegado 
            WHERE id_acceso = (
                SELECT id_acceso 
                FROM acceso_denegado 
                WHERE id_alumno = %s 
                ORDER BY fecha_intento DESC 
                LIMIT 1
            );
        """, (id_alumno,))
        
        for sol in solicitudes_restauradas:
            cursor.execute("""
                INSERT INTO historial_estados (id_solicitud, estado_anterior, estado_nuevo, id_admin, comentario)
                VALUES (%s, 'BLOQUEADA', 'APROBADA', %s, %s)
            """, (sol[0], datos.id_admin, f"Desbloqueo por apelación: {datos.comentario}"))
            
        cursor.execute("""
            UPDATE auditoria_acceso au
            SET motivo = motivo || ' [APELADO Y PERDONADO]'
            FROM asignacion a, solicitud s
            WHERE au.id_asignacion = a.id_asignacion 
              AND a.id_solicitud = s.id_solicitud
              AND s.id_alumno = %s
              AND (au.identidad_confirmada = FALSE OR au.vehiculo_coincide = FALSE)
        """, (id_alumno,))
        
        cursor.execute("""
            INSERT INTO notificaciones (numero_cuenta, rol_destino, titulo, mensaje)
            VALUES (%s, 'ALUMNO', 'Cuenta Restaurada', 'Tu apelación fue aceptada. Tu cuenta y permiso están activos nuevamente.')
        """, (numero_cuenta,))

        conexion.commit()
        correo_enviado = enviar_correo_desbloqueo(correo_electronico, nombre_completo)

        cursor.close()
        conexion.close()
        
        mensaje_final = "El alumno y sus permisos han sido desbloqueados exitosamente."
        if not correo_enviado:
            mensaje_final += " (Advertencia: Falló el envío del correo de notificación)."

        return {
            "mensaje": mensaje_final,
            "numero_cuenta": numero_cuenta,
            "permisos_restaurados": len(solicitudes_restauradas)
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.delete("/admin/usuarios/{numero_cuenta}/strikes/{id_acceso}", tags=["Administrador / Personal"], summary="Borrar un strike específico de un alumno por ID de acceso")
def borrar_strike_individual(numero_cuenta: str, id_acceso: int):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT ad.id_acceso 
            FROM acceso_denegado ad
            JOIN alumno a ON ad.id_alumno = a.id_alumno
            WHERE a.numero_cuenta = %s AND ad.id_acceso = %s;
        """, (numero_cuenta, id_acceso))

        if not cursor.fetchone():
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="El strike no existe o no corresponde al número de cuenta especificado.")

        cursor.execute("DELETE FROM acceso_denegado WHERE id_acceso = %s;", (id_acceso,))
        
        cursor.execute("""
            SELECT COUNT(*) 
            FROM acceso_denegado ad
            JOIN alumno a ON ad.id_alumno = a.id_alumno
            WHERE a.numero_cuenta = %s;
        """, (numero_cuenta,))
        strikes_restantes = cursor.fetchone()[0]

        conexion.commit()
        cursor.close()
        conexion.close()

        return {
            "mensaje": f"Strike #{id_acceso} eliminado exitosamente.",
            "numero_cuenta": numero_cuenta,
            "strikes_restantes": strikes_restantes
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.rollback()
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.get("/admin/auditoria/accesos-denegados/{numero_cuenta}", tags=["Administrador / Personal"], summary="Listar accesos denegados/strikes por alumno con paginación de 20 en 20")
def listar_accesos_denegados_alumno(numero_cuenta: str, page: int = 1):
    registros_por_pagina = 20
    offset = (page - 1) * registros_por_pagina

    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()

        cursor.execute("SELECT id_alumno, nombre, apellidos FROM alumno WHERE numero_cuenta = %s;", (numero_cuenta,))
        alumno = cursor.fetchone()
        if not alumno:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Alumno no encontrado.")

        id_alumno, nombre, apellidos = alumno[0], alumno[1], alumno[2]

        cursor.execute("SELECT COUNT(*) FROM acceso_denegado WHERE id_alumno = %s;", (id_alumno,))
        total_registros = cursor.fetchone()[0]

        cursor.execute("""
            SELECT id_acceso, motivo, fecha_intento 
            FROM acceso_denegado
            WHERE id_alumno = %s
            ORDER BY fecha_intento DESC
            LIMIT %s OFFSET %s;
        """, (id_alumno, registros_por_pagina, offset))
        filas = cursor.fetchall()

        cursor.close()
        conexion.close()

        historial = [
            {
                "id_acceso": fila[0],
                "motivo": fila[1],
                "fecha_intento": fila[2]
            }
            for fila in filas
        ]

        total_paginas = math.ceil(total_registros / registros_por_pagina) if total_registros > 0 else 1

        return {
            "numero_cuenta": numero_cuenta,
            "alumno": f"{nombre} {apellidos}",
            "pagina_actual": page,
            "registros_por_pagina": registros_por_pagina,
            "total_strikes": total_registros,
            "total_paginas": total_paginas,
            "resultados": historial
        }

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.get("/admin/guardia/auditoria-accesos", tags=["Administrador / Personal"], summary="Ver historial de accesos registrados por los guardias separado por bloques de 20 en 20")
def ver_auditoria_accesos(page: int = 1):
    registros_por_pagina = 20
    offset = (page - 1) * registros_por_pagina

    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT COUNT(*) FROM auditoria_acceso")
        total_registros = cursor.fetchone()[0]

        query = """
            SELECT 
                au.id_acceso,
                au.fecha_hora,
                au.identidad_confirmada,
                au.vehiculo_coincide,
                g.nombre || ' ' || g.apellidos AS nombre_guardia,
                al.numero_cuenta AS cuenta_alumno,
                al.nombre || ' ' || al.apellidos AS nombre_alumno,
                s.tipo_tramite,
                v.placas,
                v.modelo,
                au.motivo,
                au.evidencia_path
            FROM auditoria_acceso au
            JOIN admin g ON au.id_guardia = g.id_admin
            JOIN asignacion ag ON au.id_asignacion = ag.id_asignacion
            JOIN solicitud s ON ag.id_solicitud = s.id_solicitud
            JOIN alumno al ON s.id_alumno = al.id_alumno
            LEFT JOIN vehiculo_solicitud v ON s.id_solicitud = v.id_solicitud
            ORDER BY au.fecha_hora DESC
            LIMIT %s OFFSET %s
        """
        cursor.execute(query, (registros_por_pagina, offset))
        filas = cursor.fetchall()

        resultados = []
        for fila in filas:
            estado_acceso = "PERMITIDO" if fila[2] and fila[3] else "DENEGADO"
            
            resultados.append({
                "id_acceso": fila[0],
                "fecha_hora": fila[1],
                "estado_acceso": estado_acceso,
                "identidad_confirmada": fila[2],
                "vehiculo_coincide": fila[3],
                "guardia_turno": fila[4],
                "alumno": {
                    "numero_cuenta": fila[5],
                    "nombre": fila[6]
                },
                "tramite": fila[7].upper(),
                "vehiculo": {
                    "placas": fila[8] if fila[8] else "N/A",
                    "modelo": fila[9] if fila[9] else "N/A"
                },
                "motivo": fila[10],
                "evidencia": fila[11]
            })

        cursor.close()
        conexion.close()
        total_paginas = math.ceil(total_registros / registros_por_pagina) if total_registros > 0 else 1

        return {
            "pagina_actual": page,
            "registros_por_pagina": registros_por_pagina,
            "total_registros": total_registros,
            "total_paginas": total_paginas,
            "resultados": resultados
        }

    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.get("/admin/auditoria/evidencia-accesos-denegados/{numero_cuenta}", tags=["Administrador / Personal"], summary="Consultar evidencia de accesos denegados en caso de apelacion de un alumno")
def consultar_evidencias_apelacion(numero_cuenta: str):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")

    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT id_alumno FROM alumno WHERE numero_cuenta = %s", (numero_cuenta,))
        al_bd = cursor.fetchone()
        
        if not al_bd:
            cursor.close()
            conexion.close()
            raise HTTPException(status_code=404, detail="Alumno no encontrado")
            
        id_alumno = al_bd[0]

        cursor.execute("""
            SELECT 
                au.id_acceso,
                TO_CHAR(au.fecha_hora, 'YYYY-MM-DD HH24:MI:SS') as fecha,
                g.nombre || ' ' || g.apellidos AS nombre_guardia,
                s.tipo_tramite,
                c.folio,
                c.qr_token,
                c.vigencia,
                au.motivo,
                au.evidencia_path
            FROM auditoria_acceso au
            JOIN admin g ON au.id_guardia = g.id_admin
            JOIN asignacion ag ON au.id_asignacion = ag.id_asignacion
            JOIN solicitud s ON ag.id_solicitud = s.id_solicitud
            JOIN constancia c ON ag.id_asignacion = c.id_asignacion
            WHERE s.id_alumno = %s AND (au.identidad_confirmada = FALSE OR au.vehiculo_coincide = FALSE)
            ORDER BY au.fecha_hora DESC
        """, (id_alumno,))
        
        filas = cursor.fetchall()
        resultados = []
        for f in filas:
            resultados.append({
                "id_acceso": f[0],
                "fecha_rechazo": f[1],
                "guardia": f[2],
                "tramite": f[3].upper(),
                "folio": f[4],
                "qr_token": str(f[5]),
                "vigencia": f[6],
                "motivo": f[7],
                "evidencia": f[8]
            })

        cursor.close()
        conexion.close()
        return {"numero_cuenta": numero_cuenta, "historial_denegados": resultados}

    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")
    

@router.get("/admin/auditoria/accesos/{id_acceso}/evidencia/visualizar", tags=["Administrador / Personal"], summary="Previsualizar la evidencia fotográfica subida por el guardia de forma segura")
def ver_evidencia_guardia(id_acceso: int):
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")
        
    try:
        cursor = conexion.cursor()
        cursor.execute("SELECT evidencia_path FROM auditoria_acceso WHERE id_acceso = %s", (id_acceso,))
        resultado = cursor.fetchone()
        
        cursor.close()
        conexion.close()
        
        if not resultado or not resultado[0] or resultado[0] == 'eliminado_fin_semestre':
            raise HTTPException(status_code=404, detail="No hay evidencia fotográfica para este registro o ya fue eliminada en el cierre de semestre.")
            
        archivo_path = resultado[0]
        ruta_completa = os.path.join("uploads", archivo_path)
        
        if not os.path.exists(ruta_completa):
            raise HTTPException(status_code=404, detail="El archivo físico de la evidencia ya no existe en el servidor.")
            
        extension = archivo_path.lower().split('.')[-1]
        media_type = f"image/{extension}" if extension in ['png', 'jpg', 'jpeg'] else "application/octet-stream"
            
        return FileResponse(
            path=ruta_completa, 
            media_type=media_type, 
            filename=archivo_path,
            content_disposition_type="inline" 
        )

    except HTTPException:
        raise
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")


@router.get("/admin/estadisticas/dashboard", tags=["Administrador / Personal"], summary="Obtener estadísticas para el dashboard")
def obtener_estadisticas_dashboard():
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")
    try:
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE estado IN ('PENDIENTE', 'EN_REVISION')) as pendientes,
                COUNT(*) FILTER (WHERE estado = 'APROBADA') as aprobadas,
                COUNT(*) FILTER (WHERE estado = 'DOCUMENTACION_INCORRECTA') as rechazadas
            FROM solicitud;
        """)
        sol_stats = cursor.fetchone()
        
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE estado = 'DISPONIBLE') as disponibles,
                COUNT(*) FILTER (WHERE estado = 'OCUPADO') as ocupados,
                COUNT(*) FILTER (WHERE estado = 'MANTENIMIENTO') as mantenimiento
            FROM locker;
        """)
        lock_stats = cursor.fetchone()
        
        cursor.close()
        conexion.close()
        
        return {
            "solicitudes": {
                "total": sol_stats[0],
                "pendientes": sol_stats[1],
                "aprobadas": sol_stats[2],
                "rechazadas": sol_stats[3]
            },
            "lockers": {
                "total": lock_stats[0],
                "disponibles": lock_stats[1],
                "ocupados": lock_stats[2],
                "mantenimiento": lock_stats[3]
            }
        }
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error en BD: {str(e)}")


@router.get("/admin/auditoria/solicitudes/excel", tags=["Administrador / Personal"], summary="Descargar auditoria de todas las solicitudes en excel")
def exportar_auditoria_solicitudes_excel():
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")
        
    try:
        query = """
            SELECT 
                'FOL-' || LPAD(s.id_solicitud::text, 4, '0') as "Folio", 
                al.numero_cuenta as "Cuenta Alumno", 
                al.nombre || ' ' || al.apellidos as "Nombre Alumno", 
                al.carrera as "Carrera", 
                al.correo_electronico as "Correo Alumno", 
                s.tipo_tramite as "Trámite", 
                s.estado as "Estado Actual", 
                TO_CHAR(s.fecha_solicitud, 'YYYY-MM-DD HH24:MI') as "Fecha Solicitud",
                CASE WHEN s.estado IN ('APROBADA', 'DOCUMENTACION_INCORRECTA', 'RECHAZADA') THEN TO_CHAR(s.fecha_revision, 'YYYY-MM-DD HH24:MI') ELSE 'N/A' END as "Fecha Revisión",
                CASE WHEN s.estado IN ('APROBADA', 'DOCUMENTACION_INCORRECTA', 'RECHAZADA') THEN ad.nombre || ' ' || ad.apellidos ELSE 'N/A' END as "Revisado Por",
                CASE WHEN s.estado IN ('APROBADA', 'DOCUMENTACION_INCORRECTA', 'RECHAZADA') THEN ad.numero_cuenta ELSE 'N/A' END as "Cuenta Admin",
                CASE WHEN s.estado IN ('APROBADA', 'DOCUMENTACION_INCORRECTA', 'RECHAZADA') THEN ad.correo_electronico ELSE 'N/A' END as "Correo Admin"
            FROM solicitud s
            JOIN alumno al ON s.id_alumno = al.id_alumno
            LEFT JOIN admin ad ON s.revisado_por = ad.id_admin
            ORDER BY s.fecha_solicitud DESC
        """
        df_solicitudes = pd.read_sql(query, conexion)
        conexion.close()
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_solicitudes.to_excel(writer, sheet_name='Auditoria', startrow=5, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Auditoria']
            
            verde_uaemex = PatternFill(start_color="1D4A3C", end_color="1D4A3C", fill_type="solid")
            fuente_blanca = Font(color="FFFFFF", bold=True)
            borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            worksheet.merge_cells('C2:J3')
            celda_titulo = worksheet['C2']
            celda_titulo.value = "REGISTRO DE AUDITORÍA DE SOLICITUDES"
            celda_titulo.font = Font(size=16, bold=True, color="1D4A3C")
            celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
            
            ruta_logo = os.path.join("app", "utils", "assets", "logo_uaemex.png")
            if os.path.exists(ruta_logo):
                img = OpenpyxlImage(ruta_logo)
                img.width = 85
                img.height = 95
                worksheet.add_image(img, 'A1')
            
            for col_num, _ in enumerate(df_solicitudes.columns.values):
                celda = worksheet.cell(row=6, column=col_num + 1)
                celda.fill = verde_uaemex
                celda.font = fuente_blanca
                celda.alignment = alineacion_centro
                celda.border = borde_fino
                
            for col_num, column_cells in enumerate(worksheet.columns):
                max_length = 0
                col_letter = column_cells[0].column_letter
                
                for cell in column_cells:
                    if cell.row >= 6:
                        cell.border = borde_fino
                        if cell.row > 6:
                            cell.alignment = Alignment(vertical="center")
                    if cell.row > 5:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                ajuste = min(max_length + 3, 45)
                worksheet.column_dimensions[col_letter].width = ajuste

        output.seek(0)
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': 'attachment; filename="Registro_de_auditoria_solicitudes_CUUT.xlsx"'}, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")


@router.get("/admin/guardia/auditoria-accesos/excel", tags=["Administrador / Personal"], summary="Descargar auditoria de accesos del guardia en excel")
def exportar_auditoria_accesos_excel():
    conexion = conectar_base()
    if conexion is None:
        raise HTTPException(status_code=500, detail="Error de conexión a la BD")
        
    try:
        query = """
            SELECT 
                au.id_acceso AS "ID Acceso",
                TO_CHAR(au.fecha_hora, 'YYYY-MM-DD HH24:MI:SS') AS "Fecha y Hora",
                g.nombre || ' ' || g.apellidos AS "Guardia en Turno",
                al.numero_cuenta AS "Cuenta Alumno",
                al.nombre || ' ' || al.apellidos AS "Nombre Alumno",
                s.tipo_tramite AS "Trámite",
                CASE WHEN au.identidad_confirmada AND au.vehiculo_coincide THEN 'PERMITIDO' ELSE 'DENEGADO' END AS "Estado Acceso",
                COALESCE(au.motivo, 'N/A') AS "Motivo de Rechazo",
                CASE 
                    WHEN au.motivo LIKE '%[APELADO Y PERDONADO]%' THEN 'El usuario ha apelado su permiso'
                    WHEN al.estado_activo = FALSE AND (au.identidad_confirmada = FALSE OR au.vehiculo_coincide = FALSE) THEN 'Se suspendió su cuenta'
                    ELSE ''
                END AS "Suspensión de permiso"
            FROM auditoria_acceso au
            JOIN admin g ON au.id_guardia = g.id_admin
            JOIN asignacion ag ON au.id_asignacion = ag.id_asignacion
            JOIN solicitud s ON ag.id_solicitud = s.id_solicitud
            JOIN alumno al ON s.id_alumno = al.id_alumno
            ORDER BY au.fecha_hora DESC
        """
        df_accesos = pd.read_sql(query, conexion)
        conexion.close()
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_accesos.to_excel(writer, sheet_name='Accesos Vehiculos', startrow=5, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Accesos Vehiculos']
            
            verde_uaemex = PatternFill(start_color="1D4A3C", end_color="1D4A3C", fill_type="solid")
            rojo_suspension = Font(color="FF0000", bold=True)
            verde_apelacion = Font(color="008000", bold=True)
            fuente_blanca = Font(color="FFFFFF", bold=True)
            borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            worksheet.merge_cells('C2:H3')
            celda_titulo = worksheet['C2']
            celda_titulo.value = "REGISTRO DE AUDITORÍA DE ACCESOS VEHICULARES"
            celda_titulo.font = Font(size=15, bold=True, color="1D4A3C")
            celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
            
            ruta_logo = os.path.join("app", "utils", "assets", "logo_uaemex.png")
            if os.path.exists(ruta_logo):
                img = OpenpyxlImage(ruta_logo)
                img.width = 85
                img.height = 95
                worksheet.add_image(img, 'A1')
            
            for col_num, _ in enumerate(df_accesos.columns.values):
                celda = worksheet.cell(row=6, column=col_num + 1)
                celda.fill = verde_uaemex
                celda.font = fuente_blanca
                celda.alignment = alineacion_centro
                celda.border = borde_fino
                
            col_suspension_index = df_accesos.columns.get_loc("Suspensión de permiso") + 1
            
            for col_num, column_cells in enumerate(worksheet.columns):
                max_length = 0
                col_letter = column_cells[0].column_letter
                
                for cell in column_cells:
                    if cell.row >= 6:
                        cell.border = borde_fino
                        if cell.row > 6:
                            cell.alignment = Alignment(vertical="center")
                            if cell.column == col_suspension_index:
                                if cell.value == "Se suspendió su cuenta":
                                    cell.font = rojo_suspension
                                elif cell.value == "El usuario ha apelado su permiso":
                                    cell.font = verde_apelacion
                                    
                    if cell.row > 5:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                ajuste = min(max_length + 3, 45)
                worksheet.column_dimensions[col_letter].width = ajuste

        output.seek(0)
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': 'attachment; filename="Registro_de_auditoria_accesos_CUUT.xlsx"'}, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        if conexion:
            conexion.close()
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")